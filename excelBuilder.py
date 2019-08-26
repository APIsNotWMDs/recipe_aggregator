import json, re, openpyxl
from openpyxl.styles import Font, PatternFill

def removeWeirdSymbols(inputString):
        
    halfRegex = re.compile(r'\u00bd')
    quarterRegex = re.compile(r'\u00bc')
    threeQuarterRegex = re.compile(r'\u00be')
    thirdRegex = re.compile(r'\u2153')
    twoThirdsRegex = re.compile(r'\u2154')
    eighthRegex = re.compile(r'\u215b')
    degreeRegex = re.compile(r'\u00b0|\u00ba')
    nonBreakRegex = re.compile(r'\u00a0')
    singleQuoteRegex = re.compile(r'\u2019|\u2018')
    doubleQuoteRegex = re.compile(r'\u201c|\u201d')
    eTildeRegex = re.compile(r'\u00e9')
    eGraveRegex = re.compile(r'\u00e8')
    middleDotRegex = re.compile(r'\u00b7')
    dashRegex = re.compile(r'\u2013|\u2014')
    ellipsisRegex = re.compile(r'\u2026')
    forwardSlashRegex = re.compile(r'\u2044')

    workingString = halfRegex.sub('1/2', inputString)
    workingString = quarterRegex.sub('1/4', workingString)
    workingString = threeQuarterRegex.sub('3/4', workingString)
    workingString = thirdRegex.sub('1/3', workingString)
    workingString = twoThirdsRegex.sub('2/3', workingString)
    workingString = eighthRegex.sub('1/8', workingString)
    workingString = degreeRegex.sub(' degrees ', workingString)
    workingString = nonBreakRegex.sub('', workingString)
    workingString = doubleQuoteRegex.sub('"', workingString)
    workingString = eTildeRegex.sub('e', workingString)
    workingString = eGraveRegex.sub('e', workingString)
    workingString = forwardSlashRegex.sub('/', workingString)
    workingString = middleDotRegex.sub('', workingString)
    workingString = dashRegex.sub('-', workingString)
    workingString = ellipsisRegex.sub('...', workingString)
    outputString = singleQuoteRegex.sub('\'', workingString)

    return outputString

def main():
    
    inputFile = open("recipes.jl", 'r')

    wb = openpyxl.Workbook()
    sheet = wb.active
    titleFont = Font(bold=True, underline='single', color='ffffff')
    #lightBlueFill = PatternFill(fill_type='solid', start_color='ddebf7', end_color='ddebf7')
    darkBlueFill = PatternFill(fill_type='solid', start_color='2f75b5', end_color='2f75b5')
    
    sheet['A1'] = 'Recipe'
    sheet['A1'].font = titleFont
    sheet['A1'].fill = darkBlueFill
    sheet.column_dimensions['A'].width = 40
    sheet['B1'] = 'URL'
    sheet['B1'].font = titleFont
    sheet['B1'].fill = darkBlueFill
    sheet.column_dimensions['B'].width = 60
    sheet['C1'] = 'Ingredients'
    sheet['C1'].font = titleFont
    sheet['C1'].fill = darkBlueFill
    sheet.column_dimensions['C'].width = 65
    sheet['D1'] = 'Instructions'
    sheet['D1'].font = titleFont
    sheet['D1'].fill = darkBlueFill
    sheet.column_dimensions['D'].width = 500

    row = 2
    
    for line in inputFile:
        jsonData = json.loads(line)
        name = jsonData['Name']
        url = jsonData['URL']
        ingredients = jsonData['Ingredients']
        instructions = jsonData['Instructions']
        
        if len(ingredients) >= len(instructions):

            i = 0
            for ingredient in ingredients:
                ingredient = removeWeirdSymbols(ingredient)
                if i < len(instructions):
                    instruction = removeWeirdSymbols(instructions[i])
                    if i == 0:
                        sheet['A' + str(row)] = name
                        sheet['B' + str(row)] = url
                        sheet['C' + str(row)] = ingredient
                        sheet['D' + str(row)] = instruction
                    else:
                        sheet['C' + str(row)] = ingredient
                        sheet['D' + str(row)] = instruction
                else:
                    if i == 0:
                        sheet['A' + str(row)] = name
                        sheet['B' + str(row)] = url
                        sheet['C' + str(row)] = ingredient
                    else:
                        sheet['C' + str(row)] = ingredient

                i += 1
                row += 1

        else:

            i = 0
            for instruction in instructions:
                instruction = removeWeirdSymbols(instruction)
                if i < len(ingredients):
                    ingredient = removeWeirdSymbols(ingredients[i])
                    if i == 0:
                        sheet['A' + str(row)] = name
                        sheet['B' + str(row)] = url
                        sheet['C' + str(row)] = ingredient
                        sheet['D' + str(row)] = instruction
                    else:
                        sheet['C' + str(row)] = ingredient
                        sheet['D' + str(row)] = instruction
                else:
                    if i == 0:
                        sheet['A' + str(row)] = name
                        sheet['B' + str(row)] = url
                        sheet['D' + str(row)] = instruction
                    else:
                        sheet['D' + str(row)] = instruction

                i += 1
                row += 1

        row += 1
  
    wb.save('recipes_formatted.xlsx')
    
main()
